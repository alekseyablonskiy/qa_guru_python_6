import os
import pathlib
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook

files_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
resources_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
archive_path = os.path.join(resources_path, 'archived.zip')


# delete archive.zip
def test_archive_deleted():
    os.remove(archive_path)
    assert len(os.listdir(resources_path)) == 0


# archived files
def test_archived_files():
    with zipfile.ZipFile(archive_path, mode='w') as zip_file:
        for file in pathlib.Path(files_path).iterdir():
            zip_file.write(file, arcname=file.name)
    assert len(os.listdir(resources_path)) == 1


# read csv file
def test_read_csv():
    with zipfile.ZipFile(archive_path) as c_f:
        csv_archived = c_f.extract('csv_file.csv')
        with open(csv_archived) as csv_file:
            csv_rows = csv.reader(csv_file)
            new_list = []
            for row in csv_rows:
                new_list.append(row)
            assert len(new_list) == 5
            assert new_list == [['1', 'Python', '31%'],
                                ['2', 'Java', '17%'],
                                ['3', 'JavaScript', '8%'],
                                ['4', 'C#', '7%'],
                                ['5', 'PHP', '6%']]
        os.remove('csv_file.csv')


# read pdf file
def test_read_pdf():
    with zipfile.ZipFile(archive_path) as p_f:
        pdf_archived = p_f.extract('pdf_file.pdf')
        reader = PdfReader(pdf_archived)
        assert len(reader.pages) == 1
        page = reader.pages[0]
        text = page.extract_text()
        assert 'Hello, world!' in text
        os.remove('pdf_file.pdf')


# read xlsx file
def test_read_xlsx():
    with zipfile.ZipFile(archive_path) as x_f:
        xlsx_archived = x_f.extract('xlsx_file.xlsx')
        workbook = load_workbook(xlsx_archived)
        sheet = workbook.active
        assert sheet.cell(row=1, column=2).value == 'Aleksey'
        os.remove('xlsx_file.xlsx')
